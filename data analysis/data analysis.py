# 1. Approach: openpyxl
import openpyxl

dataframe = openpyxl.load_workbook("F00012987-WVS_Wave_7_Slovakia_Excel_v5.0.xlsx")
dataframe1 = dataframe.active
regions = []
answers = []

for col in dataframe1.iter_cols():
    if col[0].value == "N_REGION_ISO: Region ISO 3166-2":
        regions = col
    if "Q151" in col[0].value:
        answers = col

for row in range(1, dataframe1.max_row):
    print(answers[row].value)
    print(regions[row].value)
# 2. Approach: pandas
regions={'SK010': '703002', 'SK021': '703007', 'SK022': '703006', 'SK023': '703004', 'SK031': '703008','SK032': '703001', 'SK041': '703005', 'SK042': '703003'}

import pandas as pd
df = pd.read_excel('SK2022 reversed.csv')
columns_list=df.columns.to_list()
columns_list.remove('N_REGION_ISO: Region ISO 3166-2')
columns_list.remove('Q151: Willingness to fight for country')
df = df.drop(columns=columns_list)
df1=df[df["Q151: Willingness to fight for country"]==1]
grouped_main = df.groupby('N_REGION_ISO: Region ISO 3166-2').size()
grouped1 = df1.groupby('N_REGION_ISO: Region ISO 3166-2').size()
proportion1 = round((grouped1 / grouped_main)*100,2)
for region_iso,region_wvs in regions.items():
    print(proportion1[int(region_wvs)])

